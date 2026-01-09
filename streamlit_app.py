import streamlit as st
import pandas as pd
import yfinance as yf
import io
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, Optional

# --- Configuration ---
SHARIAH_THRESHOLD = 33.33
MAX_WORKERS = 8  # Increased slightly for cloud environments

# --- Core Logic Class (Refactored for Streamlit) ---
class ShariahScreener:
    def __init__(self, threshold: float = SHARIAH_THRESHOLD):
        self.threshold = threshold
        self.results = None

    def _clean_ticker(self, ticker: str) -> Optional[str]:
        """Clean and validate ticker symbols."""
        if pd.isna(ticker) or str(ticker).lower() in ['unknown', '', 'nan']:
            return None
        ticker = str(ticker).strip().upper()
        # Basic cleanup logic from your original script
        if ticker.endswith('Q.L'): ticker = ticker[:-3]
        for suffix in ['.O', '.K']:
            if ticker.endswith(suffix): ticker = ticker[:-2]
        if ticker.endswith('.S') and not ticker.endswith('.AS'):
            ticker = ticker[:-2] + '.SW'
        return ticker.replace(' ', '')

    def _get_financial_data(self, ticker: str) -> Dict:
        """Fetch data from yfinance."""
        try:
            stock = yf.Ticker(ticker)
            info = stock.info or {}
            
            # fast fail if no data
            if not info or info.get('regularMarketPrice') is None:
                return self._empty_result(ticker, "No data available")

            balance_sheet = stock.balance_sheet
            
            # --- Extract Metrics (Simplified from your script) ---
            market_cap = info.get('marketCap')
            total_debt = info.get('totalDebt')
            
            # Fallback for Debt if missing in info
            if total_debt is None and not balance_sheet.empty:
                # varied debt keys handling
                debt_keys = ['Long Term Debt', 'Current Debt', 'Total Debt'] 
                total_debt = 0
                for k in debt_keys:
                    if k in balance_sheet.index:
                        total_debt += balance_sheet.loc[k].iloc[0]

            # Cash & Securities
            cash = info.get('totalCash', 0) or 0
            # Add logic for securities if needed (simplified here for speed)
            
            # Total Assets
            total_assets = None
            if not balance_sheet.empty and 'Total Assets' in balance_sheet.index:
                total_assets = balance_sheet.loc['Total Assets'].iloc[0]

            # Receivables
            receivables = 0
            if not balance_sheet.empty:
                 for col in ['Net Receivables', 'Accounts Receivable']:
                    if col in balance_sheet.index:
                        receivables = balance_sheet.loc[col].iloc[0]
                        break

            return {
                'ticker': ticker,
                'company_name': info.get('longName', ticker),
                'sector': info.get('sector', 'N/A'),
                'market_cap': market_cap,
                'total_assets': total_assets,
                'total_debt': total_debt,
                'cash_and_securities': cash,
                'accounts_receivable': receivables,
                'data_available': True,
                'error': None
            }
        except Exception as e:
            return self._empty_result(ticker, str(e))

    def _empty_result(self, ticker, error):
        return {
            'ticker': ticker, 'company_name': ticker, 'sector': 'N/A',
            'market_cap': None, 'total_assets': None, 'total_debt': None,
            'cash_and_securities': None, 'accounts_receivable': None,
            'data_available': False, 'error': error
        }

    def _calculate_ratios(self, data: Dict) -> Dict:
        """Apply the 33% Rule."""
        res = data.copy()
        
        # Determine denominator (Market Cap vs Total Assets)
        denom = data.get('market_cap')
        if not denom or denom == 0:
            denom = data.get('total_assets')
            res['denominator_type'] = 'Total Assets'
        else:
            res['denominator_type'] = 'Market Cap'

        if not denom or denom == 0:
            res.update({'compliance_status': 'INSUFFICIENT DATA', 'overall_compliant': False})
            return res

        # Calculate
        debt = data.get('total_debt') or 0
        cash = data.get('cash_and_securities') or 0
        recv = data.get('accounts_receivable') or 0

        res['debt_ratio'] = round((debt / denom) * 100, 2)
        res['cash_ratio'] = round((cash / denom) * 100, 2)
        res['receivables_ratio'] = round((recv / denom) * 100, 2)
        
        # Check Compliance
        res['debt_compliant'] = res['debt_ratio'] < self.threshold
        res['cash_compliant'] = res['cash_ratio'] < self.threshold
        res['receivables_compliant'] = res['receivables_ratio'] < self.threshold
        
        if all([res['debt_compliant'], res['cash_compliant'], res['receivables_compliant']]):
            res['overall_compliant'] = True
            res['compliance_status'] = 'COMPLIANT'
        else:
            res['overall_compliant'] = False
            failed = []
            if not res['debt_compliant']: failed.append('Debt')
            if not res['cash_compliant']: failed.append('Cash')
            if not res['receivables_compliant']: failed.append('Recv')
            res['compliance_status'] = f"NON-COMPLIANT ({', '.join(failed)})"
            
        return res

    def screen_dataframe(self, df: pd.DataFrame, ticker_col: str):
        """Screens stocks based on a dataframe input."""
        raw_tickers = df[ticker_col].astype(str).tolist()
        tickers = [self._clean_ticker(t) for t in raw_tickers if self._clean_ticker(t)]
        tickers = list(set(tickers)) # remove dupes
        
        results = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        total = len(tickers)
        completed = 0
        
        # Run Screening in Threads
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_ticker = {executor.submit(self._screen_single, t): t for t in tickers}
            
            for future in as_completed(future_to_ticker):
                completed += 1
                try:
                    data = future.result()
                    results.append(data)
                except Exception:
                    results.append(self._empty_result(future_to_ticker[future], "Error"))
                
                # Update UI
                progress = completed / total
                progress_bar.progress(progress)
                status_text.text(f"Screening {completed}/{total}: {future_to_ticker[future]}")

        progress_bar.empty()
        status_text.empty()
        
        self.results = pd.DataFrame(results)
        return self.results

    def _screen_single(self, ticker):
        data = self._get_financial_data(ticker)
        return self._calculate_ratios(data)

    def generate_excel_bytes(self):
        """Generates the Excel file in memory."""
        output = io.BytesIO()
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='1F4E79')
        compliant_fill = PatternFill('solid', fgColor='C6EFCE')
        non_compliant_fill = PatternFill('solid', fgColor='FFC7CE')

        # Sheet 1: Results
        ws = wb.active
        ws.title = "Screening Results"
        
        headers = ['Ticker', 'Company', 'Status', 'Debt %', 'Cash %', 'Recv %', 'Market Cap']
        ws.append(headers)
        
        # Style Header
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        if self.results is not None:
            for row in self.results.itertuples():
                r = [
                    row.ticker, 
                    getattr(row, 'company_name', ''), 
                    getattr(row, 'compliance_status', ''),
                    getattr(row, 'debt_ratio', 0),
                    getattr(row, 'cash_ratio', 0),
                    getattr(row, 'receivables_ratio', 0),
                    getattr(row, 'market_cap', 0)
                ]
                ws.append(r)
                
                # Color code status
                current_row = ws.max_row
                status_cell = ws.cell(row=current_row, column=3)
                if 'COMPLIANT' in str(row.compliance_status) and 'NON' not in str(row.compliance_status):
                    status_cell.fill = compliant_fill
                else:
                    status_cell.fill = non_compliant_fill

        wb.save(output)
        output.seek(0)
        return output

# --- Streamlit App Interface ---
st.set_page_config(page_title="Shariah Screener", page_icon="☪️", layout="wide")

st.title("☪️ Shariah Compliance Screener")
st.markdown("""
This tool screens stocks based on the **Rule of 33%** (AAOIFI standards approximation):
* **Debt Ratio** < 33.33%
* **Cash Ratio** < 33.33%
* **Receivables Ratio** < 33.33%
""")

# Sidebar for options
with st.sidebar:
    st.header("Settings")
    threshold = st.number_input("Compliance Threshold (%)", value=33.33, step=0.01)
    st.info("Data provided by Yahoo Finance (yfinance).")

# File Upload
uploaded_file = st.file_uploader("Upload your CSV file (must contain a 'ticker' column)", type=["csv"])

if uploaded_file:
    df_input = pd.read_csv(uploaded_file)
    st.write("Preview of uploaded data:", df_input.head())
    
    # Attempt to find ticker column
    cols = df_input.columns
    ticker_col = next((c for c in cols if 'ticker' in c.lower() or 'symbol' in c.lower()), None)
    
    if not ticker_col:
        ticker_col = st.selectbox("Select the column containing Ticker Symbols:", cols)
    
    if st.button("Start Screening"):
        screener = ShariahScreener(threshold=threshold)
        
        with st.spinner("Fetching financial data... this may take a moment."):
            results_df = screener.screen_dataframe(df_input, ticker_col)
        
        # Display Summary Metrics
        if not results_df.empty:
            total = len(results_df)
            compliant = len(results_df[results_df['overall_compliant'] == True])
            
            c1, c2, c3 = st.columns(3)
            c1.metric("Total Screened", total)
            c2.metric("Shariah Compliant", compliant)
            c3.metric("Compliant %", f"{(compliant/total)*100:.1f}%")
            
            # Display Data
            st.subheader("Detailed Results")
            
            # visual formatting for the dataframe
            def color_status(val):
                color = '#d4edda' if val == 'COMPLIANT' else '#f8d7da' # green vs red light
                return f'background-color: {color}'
            
            st.dataframe(
                results_df[['ticker', 'compliance_status', 'debt_ratio', 'cash_ratio', 'receivables_ratio', 'company_name']],
                use_container_width=True
            )
            
            # Download Button
            excel_data = screener.generate_excel_bytes()
            st.download_button(
                label="📥 Download Excel Report",
                data=excel_data,
                file_name="shariah_compliance_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("No valid data found. Please check your ticker symbols.")
